# @File    :   ws_participant_sheet.py
# @Time    :   2019/11/15 08:51:28
# @Author  :   Wei Luo 
# @Version :   1.0
# @Contact :   luoweihoo@yahoo.com
# @Desc    :   Generate the participant sheet of a workshop

import openpyxl, sys, trfunction
from openpyxl.utils import get_column_letter

trackList = 'Team Ramp-up Tracking.xlsx'
# Check if the track list is in the current folder; if yes open it
wb = trfunction.checkTrackList(trackList)
# Build the workshop list
workshopList = trfunction.buildWorkshopList(wb)
# Ask user to enter the No. of the workshop
workshopNum, workshopCol = trfunction.chooseWorkshop(workshopList)

# Generate the participant sheet and save it to a seperate .xlsx file
participantSheet = str(workshopNum) + '_' + 'Attendance_Sheet.xlsx'
listWb = openpyxl.Workbook()
listSheet = listWb.active
listSheet['A1'] = '签到簿'
listRow = 2
sheet = wb.active
for row in range(4, sheet.max_row + 1):
    if sheet[workshopCol + str(row)].value == None:
        pass
    elif sheet[workshopCol + str(row)].value == 'X' and ( sheet['F' + str(row)].value == 'Yuedong Chen' or
            sheet['F' + str(row)].value == 'Meng Zhang' or
            sheet['F' + str(row)].value == 'Chuanyu Wang' or
            sheet['F' + str(row)].value == 'Hu Yujie' or
            sheet['F' + str(row)].value == 'Jiawei Gu' ):
        listSheet['A' + str(listRow)].value = sheet['A' + str(row)].value
        listSheet['B' + str(listRow)].value = sheet['B' + str(row)].value
        listRow = listRow + 1
listWb.save(participantSheet)
print('Processing is finished!')

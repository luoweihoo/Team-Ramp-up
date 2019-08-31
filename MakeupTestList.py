# @File    :   MakeupTestList.py
# @Time    :   2019/08/25 22:51:28
# @Author  :   Wei Luo 
# @Version :   1.0
# @Contact :   luoweihoo@yahoo.com
# @Desc    :   Generate the make-up test list of a workshop

import openpyxl, sys, trfunction
from openpyxl.utils import get_column_letter

trackList = 'Team Ramp-up Tracking.xlsx'
# Check if the track list is in the current folder; if yes open it
wb = trfunction.checkTrackList(trackList)
# Build the workshop list
workshopList = trfunction.buildWorkshopList(wb)
# Ask user to enter the No. of the workshop
workshopNum, workshopCol = trfunction.chooseWorkshop(workshopList)

# Generate the make-up test list and save it to a seperate .xlsx file
participantList = str(workshopNum) + '_' + 'MakeupList.xlsx'
listWb = openpyxl.Workbook()
listSheet = listWb.active
listRow = 1
sheet = wb.active
for row in range(4, sheet.max_row + 1):
    if sheet[workshopCol + str(row)].value == None:
        pass
    elif sheet[workshopCol + str(row)].value == 'X' or sheet[workshopCol + str(row)].value == 'X*' or int(sheet[workshopCol + str(row)].value) < 60:
        listSheet['A' + str(listRow)].value = sheet['A' + str(row)].value
        listSheet['B' + str(listRow)].value = sheet['B' + str(row)].value
        listRow = listRow + 1

listWb.save(participantList)
print('Processing is finished!')

# @File    :   trfunction.py
# @Time    :   2019/08/31 18:23:34
# @Author  :   Wei Luo
# @Version :   1.0
# @Contact :   luoweihoo@yahoo.com
# @Desc    :   Shared functions to support other programs

import openpyxl, sys
from openpyxl.utils import get_column_letter

# Check if the track list is in the current folder; if yes open it
def checkTrackList(trackList):
    print('Opening workbook...')
    try:
        wb = openpyxl.load_workbook(trackList)
        return wb
    except FileNotFoundError:
        print(f"The track list '{trackList}' doesn't exist, please check!")
        sys.exit()

# Build the workshop list from the track list
def buildWorkshopList(workbook):
    sheet = workbook.active
    workshopList = []
    colNum = 1
    for cell in sheet[2]:
        if '.' in str(cell.value):
            cellString = str(cell.value).split('.')
            colName = get_column_letter(colNum)
            cellString.append(colName)
            workshopList.append(cellString)
        colNum = colNum + 1
    return workshopList

# Ask user to choose workshop to be processed, input is the No. of the workshop e.g. 1, 2, 3...
def chooseWorkshop(workshopList):
    # User enter the workshop number
    rawInput = input("Please enter the number of the workshop!\n e.g. 1, 2..., 'q' to quit program:")
    if str(rawInput).upper() == 'Q':
        sys.exit()

    # Confirm the chosen workshop
    workshopNum = str(rawInput)
    flagFound = 'N'
    for workshop in workshopList:
        if workshopNum == workshop[0]:
            print(f"The workshop you chose is '{workshop[1].lstrip()}'.")
            rawInput = input("Please confirm the action? Y or N:")
            if str(rawInput).upper() == 'N':
                sys.exit()
            else:
                flagFound = 'Y'
                break
        # Not able to find any workshop as per the input
    if flagFound == 'N':
        print("There's no such workshop, please check!")
        sys.exit()

    # Return the column name in which the workshop is in
    return workshopNum, str(workshop[2])

# Open the test result workbook
def openTestResult(workshopNum):
    # Check wether it's a make-up test
    rawInput = input("Is this a result from a Make-up Test? Y or N:")
    if str(rawInput).upper() == 'N':
        isMakeupTest = 'N'
    else:
        isMakeupTest = 'Y'

    # Open the test result of this workshop
    if isMakeupTest == 'N':
        resultWb = workshopNum + '_' + 'Result.xlsx'
    else:
        resultWb = workshopNum + '_' + 'MakeupResult.xlsx'

    try:
        testwb = openpyxl.load_workbook(resultWb)
    except FileNotFoundError:
        print(f"The '{resultWb}' doesn't exist, please check!")
        sys.exit()

    testSheet = testwb.active
    testResult = []
    for row in range(3,testSheet.max_row + 1):
        testRow = []
        testRow.append(testSheet['C' + str(row)].value[1:])
        testRow.append(testSheet['B' + str(row)].value[1:])
        score = testSheet['F' + str(row)].value[1:]
        testRow.append(score[:-1])
        testResult.append(testRow)
    
    return testResult, isMakeupTest

# Build the summary table groupping by the manager's name
def buildSummaryList(workbook):
    groups = { 'grp1' : 'Yuedong Chen',
             'grp2' : 'Meng Zhang',
             'grp3' : 'Chuanyu Wang', 
             'grp4' : 'Pan Zhensheng'}

    sheet = workbook.active
    summaryList = []
    for k, v in groups.items():
        listRow = []
        listRow.append(v)
        listRow.append(0)       # Number of Nominated
        listRow.append(0)       # Number of quiz taken
        listRow.append(0)       # Number of passed (original test + makeup test, score >= 60)
        listRow.append(0)       # Number of failed
        summaryList.append(listRow)

#    print(summaryList)

    rowStart = 4
    colStart = 9
    for rowNum in range(rowStart, sheet.max_row + 1):
        # Check if the direct report manager is defined in our groups
        managerName = sheet['F' + str(rowNum)].value
        if  managerName in groups.values():
            for listRow in summaryList:
                # Search the record with manager's name
                if listRow[0] != managerName:
                    continue
                else:
                    for colNum in range(colStart, sheet.max_column + 1):
                        cellValue = sheet.cell(row = rowNum, column = colNum).value
                        if cellValue is None:
                            continue
                        elif cellValue == 'X*':
                            continue                
                        elif cellValue == 'X':
                            listRow[1] += 1         # Count 1 more in # of nominated
                        elif int(cellValue) >= 60:
                            listRow[1] += 1         # Count 1 more in # of nominated
                            listRow[2] += 1         # Count 1 more in # of quiz taken
                            listRow[3] += 1         # Count 1 more in # of passed
                        elif int(cellValue) < 60:
                            listRow[1] += 1         # Count 1 more in # of nominated
                            listRow[2] += 1         # Count 1 more in # of quiz taken
                            listRow[4] += 1         # Count 1 more in # of failed
                        else:
                            continue
        else:
            continue

#    print(summaryList)

    return summaryList



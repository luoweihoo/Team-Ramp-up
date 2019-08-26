# @File    :   UpdateMakeupTestResult.py
# @Time    :   2019/08/26 09:18:40
# @Author  :   Wei Luo 
# @Version :   1.0
# @Contact :   luoweihoo@yahoo.com
# @Desc    :   Update the result from the make-up test, the difference 1. read data from seperate Execl; 2. hight the score in red color

import openpyxl, sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.colors import RED

trackList = 'Team Ramp-up Tracking.xlsx'

# Check if the track list is in the current folder; if yes open it
print('Opening workbook...')
try:
    wb = openpyxl.load_workbook(trackList)
except FileNotFoundError:
    print(f"The track list '{trackList}' doesn't exist, please check!")

# Build the workshop list
sheet = wb.active
workshopList = []
colNum = 1
for cell in sheet[2]:
    if '.' in str(cell.value):
        cellString = str(cell.value).split('.')
        colName = get_column_letter(colNum)
        cellString.append(colName)
        workshopList.append(cellString)
    colNum = colNum + 1    
    
# Ask user to enter the No. of the workshop
rawInput = input("Please enter the number of the workshop!\n e.g. 1, 2..., 'q' to quit program:")
if str(rawInput).upper() == 'Q':
    sys.exit()

# Confirm the chosen workshop
workshopNum = str(rawInput)
for workshop in workshopList:
    if workshopNum == workshop[0]:
        print(f"The workshop you chose is '{workshop[1].lstrip()}'.")
        rawInput = input("Please confirm the action? Y or N:")
        if str(rawInput).upper() == 'N':
            sys.exit()
        else:
            break

# Record the column name in which the workshop is in
workshopCol = str(workshop[2])

# The chosen workshop
# print(workshopNum)
# print(workshop[1])
# print(workshop[2])

# Open the makeup test result of this workshop
resultWb = workshopNum + '_' + 'MakeupTestResult.xlsx'
try:
    testwb = openpyxl.load_workbook(resultWb)
except FileNotFoundError:
    print(f"The '{resultWb}' doesn't exist, please check!")
    sys.exit()

testSheet = testwb.active
testResult = []
for row in range(3,testSheet.max_row + 1):
    testRow = []
    testRow.append(testSheet['C' + str(row)].value[1:])
    testRow.append(testSheet['B' + str(row)].value[1:])
    score = testSheet['F' + str(row)].value[1:]
    testRow.append(score[:-1])
    testResult.append(testRow)

# print(testResult)
# print(workshopCol)

# Update the test resule into the tracking list
for result in testResult:
    for row in range(4, sheet.max_row + 1):
        if sheet['A' + str(row)].value == result[0]:
            sheet[workshopCol + str(row)].font = Font(color = RED)
            sheet[workshopCol + str(row)].value = result[2]

wb.save(trackList)
print('Processing is finished!')
